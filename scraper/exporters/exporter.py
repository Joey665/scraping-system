# ============================================================
# scraper/exporters/exporter.py
# Export scraped leads to CSV and Excel with formatting
# ============================================================

import csv
import json
from pathlib import Path
from datetime import datetime
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from scraper.core.models import EstateRecord
from scraper.utils.logger import log
from config.settings import settings

# Brand colors for Excel
BRACKEN_NAVY = "1B2A4A"
BRACKEN_GOLD = "C9A84C"
BRACKEN_LIGHT = "F5F5F5"
HOT_RED = "C0392B"
WARM_ORANGE = "E67E22"
COLD_BLUE = "2980B9"


def export_to_csv(records: list[EstateRecord], filename: Optional[str] = None) -> str:
    """Export all records to CSV. Returns file path."""
    if not records:
        log.warning("No records to export.")
        return ""

    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bracken_leads_{ts}.csv"

    path = Path(settings.OUTPUT_DIR) / filename

    fieldnames = list(records[0].to_export_dict().keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in records:
            writer.writerow(r.to_export_dict())

    log.info(f"CSV exported: {path} ({len(records):,} records)")
    return str(path)


def export_to_excel(
    records: list[EstateRecord],
    filename: Optional[str] = None,
    include_scoring_sheet: bool = True,
) -> str:
    """
    Export to formatted Excel workbook with:
    - All Leads sheet
    - HOT Prospects sheet
    - Summary Dashboard sheet
    - Conditional color coding by lead tier
    """
    if not records:
        log.warning("No records to export.")
        return ""

    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bracken_leads_{ts}.xlsx"

    path = Path(settings.OUTPUT_DIR) / filename
    wb = Workbook()

    # --- Sheet 1: All Leads ---
    ws_all = wb.active
    ws_all.title = "All Leads"
    _write_records_sheet(ws_all, records, "ALL LEADS — BRACKEN & KFEM DIGITAL SOLUTIONS")

    # --- Sheet 2: HOT Prospects ---
    hot = [r for r in records if r.lead_tier == "HOT"]
    ws_hot = wb.create_sheet("🔥 HOT Prospects")
    _write_records_sheet(ws_hot, hot, "HOT PROSPECTS — PRIORITY OUTREACH LIST")

    # --- Sheet 3: WARM Prospects ---
    warm = [r for r in records if r.lead_tier == "WARM"]
    ws_warm = wb.create_sheet("🟡 WARM Prospects")
    _write_records_sheet(ws_warm, warm, "WARM PROSPECTS")

    # --- Sheet 4: Summary Dashboard ---
    ws_dash = wb.create_sheet("📊 Dashboard", 0)
    _write_dashboard(ws_dash, records)

    wb.save(path)
    log.info(f"Excel exported: {path} ({len(records):,} records, {len(hot)} HOT)")
    return str(path)


def _write_records_sheet(ws, records: list[EstateRecord], title: str):
    """Write records to a worksheet with styling."""
    if not records:
        ws.append(["No records in this category."])
        return

    headers = list(records[0].to_export_dict().keys())

    # Title row
    ws.append([title])
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws["A1"]
    title_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor=BRACKEN_NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    ws.append(headers)
    header_row = ws[2]
    for cell in header_row:
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor=BRACKEN_GOLD)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 25

    # Data rows
    tier_colors = {"HOT": HOT_RED, "WARM": WARM_ORANGE, "COLD": COLD_BLUE}
    for i, record in enumerate(records):
        row_data = list(record.to_export_dict().values())
        ws.append(row_data)
        row_idx = i + 3

        # Zebra striping
        bg = BRACKEN_LIGHT if i % 2 == 0 else "FFFFFF"
        for cell in ws[row_idx]:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(name="Calibri", size=9)

        # Colour the Lead Tier cell
        tier = record.lead_tier
        if tier and tier in tier_colors:
            tier_col = headers.index("Lead Tier") + 1
            tier_cell = ws.cell(row=row_idx, column=tier_col)
            tier_cell.fill = PatternFill("solid", fgColor=tier_colors[tier])
            tier_cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)

    # Column widths
    col_widths = {
        "Estate Name": 35, "Google Maps URL": 45, "Address": 40,
        "State": 12, "Area / LGA": 20, "Phone": 18, "Website": 35,
        "Email": 30, "Category": 22, "Rating": 8, "Review Count": 10,
        "Opening Hours": 40, "Lead Score": 10, "Lead Tier": 10,
        "Data Quality": 12, "Scraped At": 18,
    }
    for col_idx, header in enumerate(headers, 1):
        width = col_widths.get(header, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes
    ws.freeze_panes = "A3"


def _write_dashboard(ws, records: list[EstateRecord]):
    """Write a summary dashboard to a worksheet."""
    hot = sum(1 for r in records if r.lead_tier == "HOT")
    warm = sum(1 for r in records if r.lead_tier == "WARM")
    cold = sum(1 for r in records if r.lead_tier == "COLD")
    has_phone = sum(1 for r in records if r.phone)
    has_website = sum(1 for r in records if r.website)
    lagos = sum(1 for r in records if (r.state or "").lower() == "lagos")
    abuja = sum(1 for r in records if (r.state or "").lower() in ["abuja", "fct"])
    avg_score = sum(r.lead_score or 0 for r in records) / max(len(records), 1)

    ws.title = "📊 Dashboard"

    data = [
        ["BRACKEN & KFEM DIGITAL SOLUTIONS", ""],
        ["Estate Lead Scraper — Summary Dashboard", ""],
        ["", ""],
        ["METRIC", "VALUE"],
        ["Total Records", len(records)],
        ["🔥 HOT Prospects", hot],
        ["🟡 WARM Prospects", warm],
        ["🔵 COLD Prospects", cold],
        ["", ""],
        ["Lagos Records", lagos],
        ["Abuja Records", abuja],
        ["", ""],
        ["Has Phone Number", f"{has_phone} ({has_phone/max(len(records),1)*100:.0f}%)"],
        ["Has Website", f"{has_website} ({has_website/max(len(records),1)*100:.0f}%)"],
        ["Avg Lead Score", f"{avg_score:.1f} / 100"],
        ["", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]

    for row in data:
        ws.append(row)

    # Styling
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor=BRACKEN_NAVY)
    ws["A2"].font = Font(size=11, color=BRACKEN_GOLD, name="Calibri")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    for row in ws.iter_rows(min_row=4, max_row=4):
        for cell in row:
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
            cell.fill = PatternFill("solid", fgColor=BRACKEN_GOLD)


def save_raw_json(records: list[EstateRecord], session_id: str) -> str:
    """Save raw scraped data as JSON (for re-processing without re-scraping)."""
    path = Path(settings.RAW_DATA_DIR) / f"raw_{session_id}.json"
    data = [json.loads(r.model_dump_json()) for r in records]
    path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    log.info(f"Raw JSON saved: {path}")
    return str(path)
